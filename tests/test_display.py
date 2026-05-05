"""
tests/test_display.py
====================
Tests for ui/display.py trade history rendering.
"""
from openpyxl import Workbook

from ui.display import show_trade_history


def _write_trade_headers(ws, asset_column=True):
    ws.cell(row=3, column=2, value="Date")
    if asset_column:
        ws.cell(row=3, column=3, value="Asset")
        ws.cell(row=3, column=4, value="Type")
        ws.cell(row=3, column=11, value="P&L ($)")
        ws.cell(row=3, column=14, value="Result")
        ws.cell(row=3, column=17, value="Notes")
    else:
        ws.cell(row=3, column=3, value="Type")
        ws.cell(row=3, column=10, value="P&L ($)")
        ws.cell(row=3, column=11, value="Result")
        ws.cell(row=3, column=16, value="Notes")


def _write_trade_row(ws, row, date, asset, trade_type, pnl, result="Win", notes=""):
    ws.cell(row=row, column=2, value=date)
    if asset is not None:
        ws.cell(row=row, column=3, value=asset)
        ws.cell(row=row, column=4, value=trade_type)
        ws.cell(row=row, column=11, value=pnl)
        ws.cell(row=row, column=14, value=result)
        ws.cell(row=row, column=17, value=notes)
    else:
        ws.cell(row=row, column=3, value=trade_type)
        ws.cell(row=row, column=10, value=pnl)
        ws.cell(row=row, column=11, value=result)
        ws.cell(row=row, column=16, value=notes)


def test_show_trade_history_counts_rows_with_missing_pnl(capsys):
    wb = Workbook()
    ws = wb.active
    ws.title = "📝 Paper Trades"
    _write_trade_headers(ws, asset_column=True)
    _write_trade_row(ws, 4, "2026-05-01", "ETH", "Sell Put", 150.0, "Win")
    _write_trade_row(ws, 5, "2026-05-02", "BTC", "Buy Call", None, "Open")

    show_trade_history(wb)
    out = capsys.readouterr().out

    assert "ETH" in out
    assert "BTC" in out
    assert "Total trades" in out
    assert "2" in out


def test_show_trade_history_handles_legacy_sheet_layout(capsys):
    wb = Workbook()
    ws = wb.active
    ws.title = "🔀 Strangles"
    _write_trade_headers(ws, asset_column=False)
    _write_trade_row(ws, 4, "2026-05-03", None, "Strangle", -25.0, "Loss")

    show_trade_history(wb)
    out = capsys.readouterr().out

    assert "Strangle" in out
    assert "Total trades" in out
    assert "1" in out
