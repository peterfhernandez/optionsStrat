"""
tests/test_excel_tracker.py
============================
Tests for excel/excel_tracker.py workbook migration and row append handling.
"""
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

import excel.excel_tracker as excel_tracker


def _legacy_calendar_sheet(wb):
    ws = wb.create_sheet("📅 Calendars")
    headers = [
        "Date", "Type", "Strike ($)", "Option", "Spot Open ($)", "Spot Close ($)",
        "Near Days", "Far Days", "Net Debit ($)", "P&L ($)", "Result", "Notes",
    ]
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=3, column=idx, value=header)
    return ws


def _current_calendar_sheet(wb):
    ws = wb.create_sheet("📅 Calendars")
    headers = [
        "Date", "Asset", "Type", "Strike ($)", "Option", "Spot Open ($)", "Spot Close ($)",
        "Near Days", "Far Days", "Net Debit ($)", "P&L ($)", "Result", "Notes",
    ]
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=3, column=idx, value=header)
    return ws


def test_setup_excel_migrates_legacy_calendar_sheet():
    wb = Workbook()
    wb.remove(wb.active)
    ws = _legacy_calendar_sheet(wb)
    wb.save = MagicMock()

    with patch("excel.excel_tracker.os.path.exists", return_value=True), \
         patch("excel.excel_tracker.openpyxl.load_workbook", return_value=wb):
        excel_tracker.setup_excel()

    assert ws.cell(row=3, column=3).value == "Asset"
    assert ws.cell(row=3, column=4).value == "Type"
    assert ws.cell(row=3, column=13).value == "Result"
    assert ws.cell(row=3, column=14).value == "Notes"
    wb.save.assert_called_once_with(excel_tracker.EXCEL_FILE)


def test_append_calendar_row_writes_asset_to_current_sheet_layout():
    wb = Workbook()
    wb.remove(wb.active)
    ws = _current_calendar_sheet(wb)
    wb.save = MagicMock()

    trade = {
        "date": "2026-05-06",
        "asset": "ETH",
        "type": "Call Calendar — Open (AUTO)",
        "strike": 2370,
        "option_type": "Call",
        "spot_open": 2365.91,
        "spot_close": 2365.91,
        "near_days": 7,
        "far_days": 30,
        "net_debit": 0,
        "pnl": 10.59,
        "result": "Open",
        "notes": "ETH calendar example",
    }

    excel_tracker.append_calendar_row(wb, trade)

    assert ws.cell(row=4, column=3).value == "ETH"
    assert ws.cell(row=4, column=4).value == "Call Calendar — Open (AUTO)"
    assert ws.cell(row=4, column=12).value == 10.59
    assert ws.cell(row=4, column=13).value == "Open"
    wb.save.assert_called_once_with(excel_tracker.EXCEL_FILE)
