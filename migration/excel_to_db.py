"""Migrate trade data from Excel to SQLite database."""
import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import (
    Single,
    Strangle,
    Calendar,
    get_session,
    init_db,
)


def _extract_asset_from_notes(notes: str) -> str:
    """Parse asset symbol (ETH, BTC, SOL, XRP) from notes field."""
    if not notes:
        return "ETH"  # default
    notes_upper = str(notes).upper()
    for asset in ("ETH", "BTC", "SOL", "XRP"):
        if asset in notes_upper:
            return asset
    return "ETH"  # default


def _parse_date(excel_date) -> datetime.date | None:
    """Convert Excel date (could be datetime or string) to Python date."""
    if excel_date is None:
        return None
    if isinstance(excel_date, datetime.date) and not isinstance(excel_date, datetime.datetime):
        return excel_date
    if isinstance(excel_date, datetime.datetime):
        return excel_date.date()
    if isinstance(excel_date, str):
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(excel_date, fmt).date()
            except ValueError:
                pass
    return None


def _parse_near_far_days(days_str: str) -> tuple[int | None, int | None]:
    """Parse 'near/far' days format like '7/30' into (7, 30)."""
    if not days_str:
        return None, None
    try:
        parts = str(days_str).split("/")
        near = int(parts[0].strip()) if len(parts) > 0 else None
        far = int(parts[1].strip()) if len(parts) > 1 else None
        return near, far
    except (ValueError, IndexError, AttributeError):
        return None, None


def migrate_singles(session: Session) -> int:
    """Migrate Paper Trades sheet to singles table."""
    excel_file = Path("crypto_options_trade_tracker.xlsx")
    if not excel_file.exists():
        print(f"✗ {excel_file} not found")
        return 0

    wb = load_workbook(excel_file)
    sheet_name = "📝 Paper Trades"
    if sheet_name not in wb.sheetnames:
        print(f"✗ '{sheet_name}' sheet not found in Excel")
        return 0

    ws = wb[sheet_name]
    rows_added = 0

    # Headers are in row 3; data starts at row 4
    # Columns: Col2=Date, Col3=Asset, Col4=Type, Col5=Stage, Col6=Days, Col7=Strike, Col8=SpotOpen, Col9=SpotClose, Col10=Premium, Col11=P&L, Col12=Result, Col17=Notes
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # row is a tuple; extract columns (note: col 1 is index 0, col 2 is index 1, etc.)
        if len(row) < 12 or not row[1]:  # col 2 (index 1) has the date
            continue

        # Unpack columns 2-12 and 17 (indices 1-11 and 16)
        date = row[1]  # Col 2
        asset = row[2]  # Col 3
        trade_type = row[3]  # Col 4
        stage = row[4]  # Col 5
        days = row[5]  # Col 6
        strike = row[6]  # Col 7
        spot_open = row[7]  # Col 8
        spot_close = row[8]  # Col 9
        premium = row[9]  # Col 10
        pnl = row[10]  # Col 11
        result = row[11]  # Col 12
        notes = row[16] if len(row) > 16 else None  # Col 17

        date_open = _parse_date(date)
        if not date_open:
            continue

        # Infer option_type from Type column (e.g. "Sell Cash-Secured Put" → "Put")
        option_type = None
        if trade_type:
            trade_type_str = str(trade_type).upper()
            if "PUT" in trade_type_str:
                option_type = "Put"
            elif "CALL" in trade_type_str:
                option_type = "Call"

        # Asset is already in the spreadsheet (col 3)
        asset_value = str(asset).strip() if asset else "ETH"

        trade = Single(
            asset=asset_value,
            option_type=option_type,
            stage=str(stage).strip() if stage else None,
            strike=float(strike) if strike else None,
            qty=None,  # not in Excel
            days=int(days) if days else None,
            date_open=date_open,
            date_close=None,  # Excel sheet has no close-date column
            spot_open=float(spot_open) if spot_open else None,
            spot_close=float(spot_close) if spot_close else None,
            premium=float(premium) if premium else None,
            fees=0.0,
            pnl=float(pnl) if pnl else None,
            result=str(result).strip() if result else None,
            notes=str(notes).strip() if notes else None,
        )
        session.add(trade)
        rows_added += 1

    session.commit()
    print(f"✓ Migrated {rows_added} Paper Trades → singles")
    return rows_added


def migrate_strangles(session: Session) -> int:
    """Migrate Strangles sheet to strangles table."""
    excel_file = Path("crypto_options_trade_tracker.xlsx")
    if not excel_file.exists():
        return 0

    wb = load_workbook(excel_file)
    sheet_name = "🔀 Strangles"
    if sheet_name not in wb.sheetnames:
        print(f"✗ '{sheet_name}' sheet not found in Excel")
        return 0

    ws = wb[sheet_name]
    rows_added = 0

    # Headers are in row 3; data starts at row 4
    # Columns: Col2=Date, Col3=Asset, Col4=Type, Col5=PutStrike, Col6=CallStrike, Col7=SpotOpen, Col8=SpotClose, Col9=Days, Col10=Premium, Col11=P&L, Col14=Result, Col15=Notes
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if len(row) < 12 or not row[1]:  # col 2 (index 1) has the date
            continue

        date = row[1]  # Col 2
        asset = row[2]  # Col 3
        # trade_type = row[3]  # Col 4 (not used)
        put_strike = row[4]  # Col 5
        call_strike = row[5]  # Col 6
        spot_open = row[6]  # Col 7
        spot_close = row[7]  # Col 8
        days = row[8]  # Col 9
        premium = row[9]  # Col 10
        pnl = row[10]  # Col 11
        result = row[13] if len(row) > 13 else None  # Col 14
        notes = row[14] if len(row) > 14 else None  # Col 15

        date_open = _parse_date(date)
        if not date_open:
            continue

        asset_value = str(asset).strip() if asset else "ETH"
        is_open = result and str(result).lower().strip() == "open"

        trade = Strangle(
            asset=asset_value,
            put_strike=float(put_strike) if put_strike else None,
            call_strike=float(call_strike) if call_strike else None,
            qty=None,
            days=int(days) if days else None,
            date_open=date_open,
            date_close=None,  # Excel sheet has no close-date column
            spot_open=float(spot_open) if spot_open else None,
            spot_close=float(spot_close) if spot_close else None,
            total_premium=float(premium) if premium else None,
            fees=0.0,
            pnl=float(pnl) if pnl else None,
            result=str(result).strip() if result else None,
            notes=str(notes).strip() if notes else None,
        )
        session.add(trade)
        rows_added += 1

    session.commit()
    print(f"✓ Migrated {rows_added} Strangles → strangles")
    return rows_added


def migrate_calendars(session: Session) -> int:
    """Migrate Calendars sheet to calendars table."""
    excel_file = Path("crypto_options_trade_tracker.xlsx")
    if not excel_file.exists():
        return 0

    wb = load_workbook(excel_file)
    sheet_name = "📅 Calendars"
    if sheet_name not in wb.sheetnames:
        print(f"✗ '{sheet_name}' sheet not found in Excel")
        return 0

    ws = wb[sheet_name]
    rows_added = 0

    # Headers are in row 3; data starts at row 4
    # Columns: Col2=Date, Col3=Asset, Col4=Type, Col5=Strike, Col6=Option, Col7=SpotOpen, Col8=SpotClose, Col9=NearDays, Col10=FarDays, Col11=NetDebit, Col12=P&L, Col13=Result, Col14=Notes
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if len(row) < 13 or not row[1]:  # col 2 (index 1) has the date
            continue

        date = row[1]  # Col 2
        asset = row[2]  # Col 3
        # trade_type = row[3]  # Col 4 (informational, e.g. "Put Calendar — Open")
        strike = row[4]  # Col 5
        option_type = row[5]  # Col 6
        spot_open = row[6]  # Col 7
        spot_close = row[7]  # Col 8
        near_days = row[8]  # Col 9
        far_days = row[9]  # Col 10
        net_debit = row[10]  # Col 11
        pnl = row[11]  # Col 12
        result = row[12] if len(row) > 12 else None  # Col 13
        notes = row[13] if len(row) > 13 else None  # Col 14

        date_open = _parse_date(date)
        if not date_open:
            continue

        asset_value = str(asset).strip() if asset else "ETH"
        is_open = result and str(result).lower().strip() == "open"

        trade = Calendar(
            asset=asset_value,
            option_type=str(option_type).strip() if option_type else None,
            strike=float(strike) if strike else None,
            qty=None,
            near_days=int(near_days) if near_days else None,
            far_days=int(far_days) if far_days else None,
            date_open=date_open,
            date_close=None,  # Excel sheet has no close-date column
            spot_open=float(spot_open) if spot_open else None,
            spot_close=float(spot_close) if spot_close else None,
            near_prem=None,  # not in Excel
            far_prem=None,  # not in Excel
            net_debit=float(net_debit) if net_debit else None,
            fees=0.0,
            pnl=float(pnl) if pnl else None,
            result=str(result).strip() if result else None,
            notes=str(notes).strip() if notes else None,
        )
        session.add(trade)
        rows_added += 1

    session.commit()
    print(f"✓ Migrated {rows_added} Calendars → calendars")
    return rows_added




def migrate_all() -> None:
    """Run all migrations."""
    init_db()
    session = get_session()

    try:
        singles_count = migrate_singles(session)
        strangles_count = migrate_strangles(session)
        calendars_count = migrate_calendars(session)

        total = singles_count + strangles_count + calendars_count
        print(f"\n✓ Migration complete: {total} trades migrated")
    finally:
        session.close()


if __name__ == "__main__":
    migrate_all()
