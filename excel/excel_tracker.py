"""
excel_tracker.py
================
Excel workbook setup and trade row helpers for the Crypto Options Strategy Tool.

All openpyxl logic lives here. No pricing, display, or strategy code.

Public API
----------
setup_excel()                           Open existing workbook or create a new one
append_trade_row(wb, sheet_name, trade) Append a wheel trade row to a sheet
append_strangle_row(wb, trade)          Append a strangle trade row
append_calendar_row(wb, trade)          Append a calendar spread trade row

Internal helpers
----------------
_thin_border()                          Construct a standard thin cell border
_styled(ws, ref, val, ...)              Write a value to a cell with full styling
_col_widths(ws, widths)                 Set column widths in bulk
_unmerge_row(ws, row, max_col)          Clear merged cells before writing a data row
_create_dashboard(wb)                   Build the Dashboard sheet
_create_trade_sheet(wb, name)           Build a wheel trade sheet (live or paper)
_create_strangle_sheet(wb)              Build the Strangles sheet
_create_calendar_sheet(wb)              Build the Calendar Spreads sheet
_create_summary_sheet(wb)              Build the Summary sheet

Sheet layout
------------
📊 Dashboard        KPI tiles + strategy comparison table
📋 Live Trades      Wheel live trade log
📝 Paper Trades     Wheel paper trade log
🔀 Strangles        Strangle paper trade log
📅 Calendars        Calendar spread paper trade log
📈 Summary          Cycle performance summary
"""

import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter

from config  import EXCEL_FILE, BUDGET_USD
from ui.display import ok


# ── Colour palette (Excel hex, no leading #) ──────────────────────────────────

_DARK    = "1A1A2E"
_ACCENT  = "E94560"
_GOLD    = "F5A623"
_MID     = "16213E"
_LIGHT   = "E8E8E8"
_GREEN_C = "27AE60"
_RED_C   = "E74C3C"
_WHITE_C = "FFFFFF"


# ── Low-level cell helpers ────────────────────────────────────────────────────

def _thin_border() -> Border:
    """Return a standard thin border for all four sides of a cell."""
    side = Side(style="thin", color="444444")
    return Border(left=side, right=side, top=side, bottom=side)


def _styled(
    ws,
    ref,
    val,
    bold: bool    = False,
    bg: str       = None,
    fg: str       = "000000",
    align: str    = "center",
    num_fmt: str  = None,
    italic: bool  = False,
    size: int     = 10,
):
    """
    Write a value to a cell and apply full styling in one call.

    Parameters
    ----------
    ws      : Worksheet
    ref     : str | Cell   Cell reference (e.g. "B2") or Cell object
    val     : any          Value to write
    bold    : bool         Bold font
    bg      : str | None   Background fill colour (hex, no #)
    fg      : str          Font colour (hex, no #)
    align   : str          Horizontal alignment
    num_fmt : str | None   Number format string (e.g. "$#,##0.00")
    italic  : bool         Italic font
    size    : int          Font size
    """
    cell = ws[ref] if isinstance(ref, str) else ref
    cell.value     = val
    cell.font      = Font(name="Arial", bold=bold, italic=italic, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _col_widths(ws, widths: list[int]) -> None:
    """Set column widths from a list, starting at column A."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _unmerge_row(ws, row: int, max_col: int = 17) -> None:
    """
    Unmerge any merged cells in a row and clear their values.

    Required before writing data into a row that previously held
    a merged placeholder cell (e.g. the "← trades appear here" banner).
    """
    merges_to_remove = [
        rng for rng in ws.merged_cells.ranges
        if rng.min_row <= row <= rng.max_row
    ]
    for rng in merges_to_remove:
        ws.unmerge_cells(str(rng))
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).value = None


# ── Sheet builders ────────────────────────────────────────────────────────────

def _create_dashboard(wb) -> None:
    """Build the 📊 Dashboard sheet with KPI tiles and strategy comparison."""
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3, 32, 22, 22, 22])

    # Title
    ws.row_dimensions[2].height = 45
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value     = "ETH OPTIONS STRATEGY TRACKER v3"
    c.font      = Font(name="Arial", bold=True, color=_ACCENT, size=20)
    c.fill      = PatternFill("solid", start_color=_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI tiles
    kpis = [
        (4,  "Budget (USD)",                    "250",                          "$#,##0.00", _GOLD),
        (5,  "Wheel — Total Paper Premium",      "='📝 Paper Trades'!K2",        "$#,##0.00", _GREEN_C),
        (6,  "Wheel — Paper Win Rate",           "='📝 Paper Trades'!L2",        "0.0%",      _GOLD),
        (7,  "Wheel — Paper Cycles",             "='📝 Paper Trades'!M2",        "0",         _LIGHT),
        (8,  "Strangle — Total Paper Premium",   "='🔀 Strangles'!H2",           "$#,##0.00", _GREEN_C),
        (9,  "Strangle — Paper Win Rate",        "='🔀 Strangles'!I2",           "0.0%",      _GOLD),
        (10, "Strangle — Trades Completed",      "='🔀 Strangles'!J2",           "0",         _LIGHT),
        (11, "Calendar — Total P&L",             "='📅 Calendars'!H2",           "$#,##0.00", _GREEN_C),
        (12, "Calendar — Win Rate",              "='📅 Calendars'!I2",           "0.0%",      _GOLD),
        (13, "Calendar — Trades Completed",      "='📅 Calendars'!J2",           "0",         _LIGHT),
        (14, "Live — Total Premium",             "='📋 Live Trades'!K2",         "$#,##0.00", _GREEN_C),
        (15, "Live — Win Rate",                  "='📋 Live Trades'!L2",         "0.0%",      _GOLD),
    ]
    for row, label, formula, fmt, color in kpis:
        ws.row_dimensions[row].height = 24
        _styled(ws, f"B{row}", label,    bold=True, bg=_MID,  fg=_LIGHT, align="left")
        _styled(ws, f"C{row}", formula,             bg=_DARK, fg=color,  num_fmt=fmt)

    # Disclaimer
    ws.row_dimensions[17].height = 18
    ws.merge_cells("B17:E17")
    c = ws["B17"]
    c.value     = "ℹ Premiums are Black-Scholes estimates. Paper trading only — not financial advice."
    c.font      = Font(name="Arial", italic=True, color="888888", size=9)
    c.alignment = Alignment(horizontal="center")

    # Strategy comparison table
    ws.row_dimensions[19].height = 22
    ws.merge_cells("B19:E19")
    _styled(ws, "B19", "STRATEGY COMPARISON", bold=True, bg=_MID, fg=_GOLD, size=11)

    for col, header in enumerate(["Strategy", "Max Income/Loss", "Risk", "Best Market"], 2):
        _styled(ws, f"{get_column_letter(col)}20", header, bold=True, bg=_DARK, fg=_ACCENT)

    rows_data = [
        ("Wheel (CSP→CC)",    "Medium premium",    "Capped (put side)", "Mild bullish / sideways"),
        ("Short Strangle",    "High premium",      "Unlimited",         "Sideways / low vol"),
        ("Calendar Spread",   "Limited profit",    "Capped ✓ (debit)",  "Neutral / low movement"),
        ("Short Straddle",    "Highest premium",   "Unlimited",         "Very sideways"),
        ("Iron Condor",       "Low premium",       "Capped ✓",          "Range-bound"),
    ]
    for i, row_data in enumerate(rows_data, 21):
        ws.row_dimensions[i].height = 20
        bg = _MID if i % 2 == 0 else _DARK
        for col, val in enumerate(row_data, 2):
            _styled(ws, f"{get_column_letter(col)}{i}", val, bg=bg, fg=_LIGHT, align="left")


def _create_trade_sheet(wb, name: str) -> None:
    """
    Build a wheel trade log sheet (used for both Live and Paper Trades).

    Hidden summary formulas in row 2 feed the Dashboard KPI tiles.
    """
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3, 7, 13, 22, 18, 10, 13, 14, 14, 13, 13, 10, 18, 12, 12, 12, 28])

    # Hidden summary formulas (read by Dashboard)
    ws.row_dimensions[2].height = 1
    ws["K2"] = "=IF(COUNTA(K4:K5000)=0,0,SUM(K4:K5000))"
    ws["L2"] = '=IF(COUNTA(L4:L5000)=0,0,COUNTIF(L4:L5000,"Win")/COUNTA(L4:L5000))'
    ws["M2"] = '=COUNTIF(D4:D5000,"Call Expired Worthless")'
    ws["K2"].number_format = "$#,##0.00"
    ws["L2"].number_format = "0.0%"

    # Column headers
    ws.row_dimensions[3].height = 34
    headers = [
        "Date", "Asset", "Type", "Stage", "Days", "Strike ($)", "Spot Open ($)",
        "Spot Close ($)", "Premium ($)", "P&L ($)", "Result",
        "Cum. Premium ($)", "Win Rate", "Cycles", "", "Notes",
    ]
    for col, header in enumerate(headers, 2):
        ref     = f"{get_column_letter(col)}3"
        ws[ref] = header
        ws[ref].font      = Font(name="Arial", bold=True, color=_ACCENT, size=10)
        ws[ref].fill      = PatternFill("solid", start_color=_DARK)
        ws[ref].border    = _thin_border()
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Placeholder row
    ws.row_dimensions[4].height = 18
    ws.merge_cells("B4:Q4")
    c = ws["B4"]
    c.value     = "← Trades will appear here as you use the tool"
    c.font      = Font(name="Arial", italic=True, color="666666", size=9)
    c.fill      = PatternFill("solid", start_color=_MID)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _create_strangle_sheet(wb) -> None:
    """
    Build the 🔀 Strangles sheet for strangle paper trade logs.

    Hidden summary formulas in row 2 feed the Dashboard KPI tiles.
    """
    ws = wb.create_sheet("🔀 Strangles")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3, 7, 13, 14, 14, 14, 14, 14, 14, 12, 12, 16, 16, 28])

    # Hidden summary formulas (read by Dashboard)
    ws.row_dimensions[2].height = 1
    ws["H2"] = "=IF(COUNTA(I4:I5000)=0,0,SUM(I4:I5000))"
    ws["I2"] = '=IF(COUNTA(L4:L5000)=0,0,COUNTIF(L4:L5000,"Win")/COUNTA(L4:L5000))'
    ws["J2"] = "=COUNTA(A4:A5000)"
    ws["H2"].number_format = "$#,##0.00"
    ws["I2"].number_format = "0.0%"

    # Column headers
    ws.row_dimensions[3].height = 34
    headers = [
        "Date", "Asset", "Type", "Put Strike", "Call Strike", "Spot Open", "Spot Close",
        "Days", "Premium ($)", "P&L ($)", "Lower B/E", "Upper B/E", "Result", "Notes",
    ]
    for col, header in enumerate(headers, 2):
        ref     = f"{get_column_letter(col)}3"
        ws[ref] = header
        ws[ref].font      = Font(name="Arial", bold=True, color=_ACCENT, size=10)
        ws[ref].fill      = PatternFill("solid", start_color=_DARK)
        ws[ref].border    = _thin_border()
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Placeholder row
    ws.row_dimensions[4].height = 18
    ws.merge_cells("B4:O4")
    c = ws["B4"]
    c.value     = "← Strangle trades will appear here as you paper trade"
    c.font      = Font(name="Arial", italic=True, color="666666", size=9)
    c.fill      = PatternFill("solid", start_color=_MID)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _create_calendar_sheet(wb) -> None:
    """
    Build the 📅 Calendars sheet for calendar spread paper trade logs.

    Columns: Date, Type, Strike, Option Type, Spot Open, Spot Close,
             Near Days, Far Days, Net Debit, P&L, Result, Notes.
    Hidden summary formulas in row 2 feed the Dashboard KPI tiles.
    """
    ws = wb.create_sheet("📅 Calendars")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3, 7, 13, 20, 12, 12, 14, 14, 11, 10, 13, 13, 12, 28])

    # Hidden summary formulas (read by Dashboard)
    ws.row_dimensions[2].height = 1
    ws["H2"] = "=IF(COUNTA(K4:K5000)=0,0,SUM(K4:K5000))"   # total P&L
    ws["I2"] = '=IF(COUNTA(L4:L5000)=0,0,COUNTIF(L4:L5000,"Win")/COUNTA(L4:L5000))'
    ws["J2"] = "=COUNTA(A4:A5000)"                          # trade count
    ws["H2"].number_format = "$#,##0.00"
    ws["I2"].number_format = "0.0%"

    # Column headers
    ws.row_dimensions[3].height = 34
    headers = [
        "Date", "Asset", "Type", "Strike ($)", "Option", "Spot Open ($)", "Spot Close ($)",
        "Near Days", "Far Days", "Net Debit ($)", "P&L ($)", "Result", "Notes",
    ]
    for col, header in enumerate(headers, 2):
        ref     = f"{get_column_letter(col)}3"
        ws[ref] = header
        ws[ref].font      = Font(name="Arial", bold=True, color=_ACCENT, size=10)
        ws[ref].fill      = PatternFill("solid", start_color=_DARK)
        ws[ref].border    = _thin_border()
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Placeholder row
    ws.row_dimensions[4].height = 18
    ws.merge_cells("B4:O4")
    c = ws["B4"]
    c.value     = "← Calendar spread trades will appear here as you paper trade"
    c.font      = Font(name="Arial", italic=True, color="666666", size=9)
    c.fill      = PatternFill("solid", start_color=_MID)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _headers(ws):
    """Return a mapping of header text to column index for row 3."""
    return {
        str(ws.cell(row=3, column=col).value).strip(): col
        for col in range(2, ws.max_column + 1)
        if ws.cell(row=3, column=col).value is not None
    }


def _ensure_asset_column(ws) -> bool:
    """Insert an Asset column for legacy sheets that still use old header layout."""
    headers = _headers(ws)
    if "Asset" in headers:
        return False

    # Legacy sheet layout: Date, Type, ... with no Asset column.
    if headers.get("Type") == 3:
        ws.insert_cols(3)
        ws.cell(row=3, column=3, value="Asset")
        return True

    return False


def _create_summary_sheet(wb) -> None:
    """Build the 📈 Summary sheet header (rows are populated at runtime)."""
    ws = wb.create_sheet("📈 Summary")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3, 18, 16, 16, 16, 16, 30])

    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value     = "CYCLE PERFORMANCE SUMMARY"
    c.font      = Font(name="Arial", bold=True, color=_ACCENT, size=16)
    c.fill      = PatternFill("solid", start_color=_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")


# ── Public API ────────────────────────────────────────────────────────────────

def setup_excel():
    """
    Open the existing workbook or create a fresh one with all sheets.

    If the workbook already exists but is missing the Strangles sheet
    (e.g. upgrading from an older version), the sheet is added automatically.

    Returns
    -------
    openpyxl.Workbook
    """
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        changed = False
        if "🔀 Strangles" not in wb.sheetnames:
            _create_strangle_sheet(wb)
            changed = True
        if "📅 Calendars" not in wb.sheetnames:
            _create_calendar_sheet(wb)
            changed = True

        for sheet_name in ["📋 Live Trades", "📝 Paper Trades", "🔀 Strangles", "📅 Calendars"]:
            if sheet_name in wb.sheetnames and _ensure_asset_column(wb[sheet_name]):
                changed = True

        if changed:
            wb.save(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)          # remove default empty sheet
        _create_dashboard(wb)
        _create_trade_sheet(wb, "📋 Live Trades")
        _create_trade_sheet(wb, "📝 Paper Trades")
        _create_strangle_sheet(wb)
        _create_calendar_sheet(wb)
        _create_summary_sheet(wb)
        wb.save(EXCEL_FILE)
    return wb


def append_trade_row(wb, sheet_name: str, trade: dict) -> None:
    """
    Append a wheel trade row to a trade sheet.

    Finds the first empty row after the headers (or replaces the placeholder
    row if no trades have been recorded yet), writes the trade data with
    colour-coded styling based on result, and saves the workbook.

    Parameters
    ----------
    wb         : openpyxl.Workbook
    sheet_name : str   "📋 Live Trades" or "📝 Paper Trades"
    trade      : dict  Keys: date, type, stage, days, strike, spot_open,
                       spot_close, premium, pnl, result, notes
    """
    ws  = wb[sheet_name]
    row = 4
    while True:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            break
        if isinstance(cell_val, str) and "←" in cell_val:
            _unmerge_row(ws, row)
            break
        row += 1

    headers = _headers(ws)
    has_asset = "Asset" in headers
    result = trade.get("result", "")
    bg = "1a3a1a" if result == "Win" else "3a1a1a" if result == "Loss" else _MID
    fg = _GREEN_C  if result == "Win" else _RED_C   if result == "Loss" else _LIGHT

    if has_asset:
        values = [
            trade.get("date",       ""),
            trade.get("asset",      ""),
            trade.get("type",       ""),
            trade.get("stage",      ""),
            trade.get("days",       7),
            trade.get("strike",     0),
            trade.get("spot_open",  0),
            trade.get("spot_close", ""),
            trade.get("premium",    0),
            trade.get("pnl",        ""),
            result,
            f"=SUM($K$4:K{row})",
            "", "", "",
            trade.get("notes", ""),
        ]
    else:
        values = [
            trade.get("date",       ""),
            trade.get("type",       ""),
            trade.get("stage",      ""),
            trade.get("days",       7),
            trade.get("strike",     0),
            trade.get("spot_open",  0),
            trade.get("spot_close", ""),
            trade.get("premium",    0),
            trade.get("pnl",        ""),
            result,
            f"=SUM($J$4:J{row})",
            "", "", "",
            trade.get("notes", ""),
        ]

    fmts = {7: "$#,##0", 8: "$#,##0.00", 9: "$#,##0.00",
            10: "$#,##0.00", 12: "$#,##0.00"}

    for col_idx, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font      = Font(name="Arial", size=9, color=fg)
        c.fill      = PatternFill("solid", start_color=bg)
        c.border    = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in fmts:
            c.number_format = fmts[col_idx]

    ws.row_dimensions[row].height = 18
    wb.save(EXCEL_FILE)
    ok(f"Trade saved → {sheet_name} row {row}")


def append_strangle_row(wb, trade: dict) -> None:
    """
    Append a strangle trade row to the 🔀 Strangles sheet.

    Automatically computes breakeven prices from the trade data.
    Colour-codes the row green (Win), red (Loss), or neutral (Open).

    Parameters
    ----------
    wb    : openpyxl.Workbook
    trade : dict  Keys: date, type, put_strike, call_strike, spot_open,
                  spot_close, days, premium, pnl, result, notes
    """
    ws  = wb["🔀 Strangles"]
    row = 4
    while True:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            break
        if isinstance(cell_val, str) and "←" in cell_val:
            _unmerge_row(ws, row, max_col=15)
            break
        row += 1

    headers = _headers(ws)
    has_asset = "Asset" in headers
    result = trade.get("result", "")
    bg = "1a3a1a" if result == "Win" else "3a1a1a" if result == "Loss" else _MID
    fg = _GREEN_C  if result == "Win" else _RED_C   if result == "Loss" else _LIGHT

    # Compute breakevens from trade data
    Kp            = trade.get("put_strike",  0)
    Kc            = trade.get("call_strike", 0)
    prem          = trade.get("premium",     0)
    spot_open     = trade.get("spot_open",   1) or 1
    qty           = BUDGET_USD / spot_open
    prem_per_unit = prem / qty if qty else 0
    be_lo         = Kp - prem_per_unit
    be_hi         = Kc + prem_per_unit

    if has_asset:
        values = [
            trade.get("date",       ""),
            trade.get("asset",      ""),
            trade.get("type",       ""),
            Kp,
            Kc,
            trade.get("spot_open",  ""),
            trade.get("spot_close", ""),
            trade.get("days",       7),
            prem,
            trade.get("pnl",        ""),
            be_lo,
            be_hi,
            result,
            trade.get("notes",      ""),
        ]
    else:
        values = [
            trade.get("date",       ""),
            trade.get("type",       ""),
            Kp,
            Kc,
            trade.get("spot_open",  ""),
            trade.get("spot_close", ""),
            trade.get("days",       7),
            prem,
            trade.get("pnl",        ""),
            be_lo,
            be_hi,
            result,
            trade.get("notes",      ""),
        ]

    fmts = {
        5: "$#,##0", 6: "$#,##0", 7: "$#,##0.00", 8: "$#,##0.00",
        10: "$#,##0.00", 11: "$#,##0.00", 12: "$#,##0.00", 13: "$#,##0.00",
    }

    for col_idx, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font      = Font(name="Arial", size=9, color=fg)
        c.fill      = PatternFill("solid", start_color=bg)
        c.border    = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in fmts:
            c.number_format = fmts[col_idx]

    ws.row_dimensions[row].height = 18
    wb.save(EXCEL_FILE)
    ok(f"Strangle trade saved → 🔀 Strangles row {row}")


def append_calendar_row(wb, trade: dict) -> None:
    """
    Append a calendar spread trade row to the 📅 Calendars sheet.

    Colour-codes the row green (Win), red (Loss), or neutral (Open).

    Parameters
    ----------
    wb    : openpyxl.Workbook
    trade : dict  Keys: date, type, strike, option_type, spot_open, spot_close,
                  near_prem, far_prem, net_debit, pnl, near_days, far_days,
                  result, notes
    """
    ws  = wb["📅 Calendars"]
    row = 4
    while True:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            break
        if isinstance(cell_val, str) and "←" in cell_val:
            _unmerge_row(ws, row, max_col=15)
            break
        row += 1

    headers = _headers(ws)
    has_asset = "Asset" in headers
    result = trade.get("result", "")
    bg = "1a3a1a" if result == "Win" else "3a1a1a" if "Loss" in result else _MID
    fg = _GREEN_C  if result == "Win" else _RED_C   if "Loss" in result else _LIGHT

    if has_asset:
        values = [
            trade.get("date",        ""),
            trade.get("asset",       ""),
            trade.get("type",        ""),
            trade.get("strike",      0),
            trade.get("option_type", ""),
            trade.get("spot_open",   ""),
            trade.get("spot_close",  ""),
            trade.get("near_days",   7),
            trade.get("far_days",    30),
            trade.get("net_debit",   0),
            trade.get("pnl",         ""),
            result,
            trade.get("notes",       ""),
        ]
    else:
        values = [
            trade.get("date",        ""),
            trade.get("type",        ""),
            trade.get("strike",      0),
            trade.get("option_type", ""),
            trade.get("spot_open",   ""),
            trade.get("spot_close",  ""),
            trade.get("near_days",   7),
            trade.get("far_days",    30),
            trade.get("net_debit",   0),
            trade.get("pnl",         ""),
            result,
            trade.get("notes",       ""),
        ]

    fmts = {
        5: "$#,##0", 7: "$#,##0.00", 8: "$#,##0.00",
        11: "$#,##0.00", 12: "$#,##0.00",
    }

    for col_idx, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font      = Font(name="Arial", size=9, color=fg)
        c.fill      = PatternFill("solid", start_color=bg)
        c.border    = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in fmts:
            c.number_format = fmts[col_idx]

    ws.row_dimensions[row].height = 18
    wb.save(EXCEL_FILE)
    ok(f"Calendar trade saved → 📅 Calendars row {row}")
