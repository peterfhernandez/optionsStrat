"""
Crypto Options Strategy Tool  v3.0
=================================
Paper trading tool for two strategies on ETH options:
  1. Wheel Strategy  (Cash-Secured Put → Covered Call)
  2. Short Strangle  (Sell OTM Put + OTM Call simultaneously)

New in v3:
  - Short Strangle paper trading simulator
  - ASCII profit zone visualiser (breakevens, profit/loss at expiry)
  - Strangle analysis table (strike pairs, combined premium, breakevens)
  - Strangle trades tracked in Excel (dedicated sheet)
  - Combined stats dashboard

Requirements:
  pip install requests openpyxl
"""

import json, math, os, shutil
from datetime import datetime, timedelta, date

try:
    import requests
except ImportError:
    print("Please run: pip install requests openpyxl"); exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Please run: pip install openpyxl"); exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
BUDGET_USD     = 250.0
EXCEL_FILE     = "crypto_options_trade_tracker.xlsx"
RISK_FREE_RATE = 0.05
OTM_LEVELS     = [0.10, 0.15, 0.20]

# Stop-loss: close the strangle if its current value reaches this multiple of
# the original premium received.  2.0 = "buy back at 2x what you sold it for"
# (i.e. you've lost an amount equal to your original premium).
# Common professional rule: 2x stop = max loss equals 1x the premium collected.
STOP_LOSS_MULTIPLIER = 2.0   # ← change this to adjust your stop level

# Warning level: flag the position (but don't force close) at this multiple
STOP_WARN_MULTIPLIER = 1.5   # ← yellow alert at 1.5x current value

# ── ANSI colours ──────────────────────────────────────────────────────────────
R="\033[0m"; B="\033[1m"; CY="\033[96m"; GR="\033[92m"
YL="\033[93m"; RD="\033[91m"; GY="\033[90m"; WH="\033[97m"
BG_GR="\033[42m"; BG_RD="\033[41m"; BG_YL="\033[43m"

def hdr(t):
    print(f"\n{CY}{'─'*60}{R}\n{B}{CY}  {t}{R}\n{CY}{'─'*60}{R}")

def sub(t):   print(f"\n{YL}  ▸ {t}{R}")
def inf(l,v): print(f"    {GY}{l:<34}{R}{WH}{v}{R}")
def ok(t):    print(f"  {GR}✓ {t}{R}")
def warn(t):  print(f"  {YL}⚠ {t}{R}")
def err(t):   print(f"  {RD}✗ {t}{R}")

# ── Market data ───────────────────────────────────────────────────────────────

def get_eth_price():
    try:
        r = requests.get("https://api.coingecko.com/api/v3/simple/price",
                         params={"ids":"ethereum","vs_currencies":"usd"}, timeout=8)
        return float(r.json()["ethereum"]["usd"])
    except Exception as e:
        warn(f"ETH price fetch failed: {e}"); return None


def get_deribit_iv(spot, days):
    try:
        today = datetime.utcnow()
        if days == 1:
            expiry_date = today + timedelta(days=1)
        else:
            days_until_friday = (4 - today.weekday()) % 7 or 7
            expiry_date = today + timedelta(days=days_until_friday)
        expiry_str = expiry_date.strftime("%d%b%y").upper()
        atm_strike = round(spot / 100) * 100
        for opt in ["P","C"]:
            inst = f"ETH-{expiry_str}-{int(atm_strike)}-{opt}"
            r = requests.get("https://www.deribit.com/api/v2/public/get_order_book",
                             params={"instrument_name": inst}, timeout=8)
            if r.status_code == 200:
                iv = r.json().get("result",{}).get("mark_iv")
                if iv and iv > 0:
                    return float(iv) / 100.0
    except Exception as e:
        warn(f"IV fetch failed: {e}")
    return None

# ── Black-Scholes ─────────────────────────────────────────────────────────────

def ncdf(x):
    return 0.5 * math.erfc(-x / math.sqrt(2))

def bs_put(S, K, T, r, v):
    if T <= 0 or v <= 0: return max(K-S, 0)
    d1 = (math.log(S/K) + (r + 0.5*v**2)*T) / (v*math.sqrt(T))
    d2 = d1 - v*math.sqrt(T)
    return K*math.exp(-r*T)*ncdf(-d2) - S*ncdf(-d1)

def bs_call(S, K, T, r, v):
    if T <= 0 or v <= 0: return max(S-K, 0)
    d1 = (math.log(S/K) + (r + 0.5*v**2)*T) / (v*math.sqrt(T))
    d2 = d1 - v*math.sqrt(T)
    return S*ncdf(d1) - K*math.exp(-r*T)*ncdf(d2)

def prob_otm_put(S, K, T, r, v):
    if T <= 0 or v <= 0: return 1.0 if S > K else 0.0
    d2 = (math.log(S/K) + (r - 0.5*v**2)*T) / (v*math.sqrt(T))
    return ncdf(d2)

def prob_otm_call(S, K, T, r, v):
    if T <= 0 or v <= 0: return 1.0 if S < K else 0.0
    d2 = (math.log(S/K) + (r - 0.5*v**2)*T) / (v*math.sqrt(T))
    return ncdf(-d2)

# ── Strangle helpers ──────────────────────────────────────────────────────────

def strangle_pnl(spot_at_expiry, put_strike, call_strike, total_premium, qty):
    """P&L of short strangle at expiry for given ETH price."""
    put_loss  = max(put_strike  - spot_at_expiry, 0) * qty
    call_loss = max(spot_at_expiry - call_strike, 0) * qty
    return total_premium - put_loss - call_loss

def strangle_breakevens(put_strike, call_strike, premium_per_eth):
    """Lower and upper breakeven prices."""
    lower = put_strike  - premium_per_eth
    upper = call_strike + premium_per_eth
    return lower, upper

# ── Stop-loss checker ────────────────────────────────────────────────────────

def check_stop_loss(spot, iv, days, op):
    """
    Evaluate stop-loss status for an open strangle.
    Returns: ("ok" | "warn" | "stop"), current_value, multiplier, message
    """
    T       = max(days / 365.0, 1/365.0)
    q       = op["qty"]
    p0      = op["total_premium"]
    Kp      = op["put_strike"]
    Kc      = op["call_strike"]
    cur_pp  = bs_put (spot, Kp, T, RISK_FREE_RATE, iv) * q
    cur_cp  = bs_call(spot, Kc, T, RISK_FREE_RATE, iv) * q
    cur_val = cur_pp + cur_cp
    mult    = cur_val / p0 if p0 > 0 else 0

    if mult >= STOP_LOSS_MULTIPLIER:
        pnl = p0 - cur_val
        msg = (f"STOP-LOSS TRIGGERED  strangle now worth ${cur_val:.2f} "
               f"({mult:.1f}x premium).  Est. loss if closed now: ${pnl:.2f}")
        return "stop", cur_val, mult, msg
    elif mult >= STOP_WARN_MULTIPLIER:
        msg = (f"Stop WARNING  strangle now worth ${cur_val:.2f} "
               f"({mult:.1f}x premium).  Stop triggers at {STOP_LOSS_MULTIPLIER:.1f}x "
               f"(${p0*STOP_LOSS_MULTIPLIER:.2f}).")
        return "warn", cur_val, mult, msg
    else:
        msg = (f"OK  {mult:.2f}x premium  "
               f"(warn at {STOP_WARN_MULTIPLIER:.1f}x, stop at {STOP_LOSS_MULTIPLIER:.1f}x)")
        return "ok", cur_val, mult, msg


def print_stop_loss_status(status, cur_val, mult, msg, p0):
    """Print a formatted stop-loss status bar."""
    bar_width = 36
    fill_pct  = min(mult / 3.0, 1.0)
    stop_pos  = int((STOP_LOSS_MULTIPLIER / 3.0) * bar_width)
    warn_pos  = int((STOP_WARN_MULTIPLIER  / 3.0) * bar_width)
    filled    = int(fill_pct * bar_width)

    bar = ""
    for i in range(bar_width):
        if i < filled:
            if i >= stop_pos:   bar += f"{RD}█{R}"
            elif i >= warn_pos: bar += f"{YL}█{R}"
            else:               bar += f"{GR}█{R}"
        elif i == stop_pos:     bar += f"{RD}|{R}"
        elif i == warn_pos:     bar += f"{YL}|{R}"
        else:                   bar += f"{GY}░{R}"

    print(f"\n  {WH}Stop-Loss Monitor{R}  "
          f"{GY}[warn={STOP_WARN_MULTIPLIER:.1f}x  stop={STOP_LOSS_MULTIPLIER:.1f}x]{R}")
    print(f"  {bar}  {WH}{mult:.2f}x premium{R}")

    if status == "stop":
        print(f"\n  {RD}{'━'*54}{R}")
        print(f"  {RD}  ⛔  {msg}{R}")
        print(f"  {RD}{'━'*54}{R}")
        print(f"\n  {WH}Recommended action:{R} Close the position now (option {CY}[3]{R}).")
        print(f"  {WH}Why:{R} Closing at {STOP_LOSS_MULTIPLIER:.0f}x caps your loss at ~1x your premium.")
        print(f"       Holding exposes you to much larger losses if ETH keeps moving.\n")
    elif status == "warn":
        print(f"\n  {YL}  ⚠  {msg}{R}")
        print(f"  {GY}  Plan your exit. Hard stop triggers at ${p0*STOP_LOSS_MULTIPLIER:.2f}.\n{R}")
    else:
        print(f"  {GY}  ✓  {msg}{R}\n")


# ── ASCII Profit Zone Visualiser ──────────────────────────────────────────────

def draw_profit_zone(spot, put_strike, call_strike, total_premium, qty, days, iv):
    """
    Draws an ASCII chart showing P&L across a price range at expiry.
    Also overlays a simple probability distribution hint.
    """
    width   = shutil.get_terminal_size((80,20)).columns - 4
    width   = max(60, min(width, 100))

    prem_per_eth = total_premium / qty if qty > 0 else 0
    be_lo, be_hi = strangle_breakevens(put_strike, call_strike, prem_per_eth)

    # Price range: spot ± 35%
    lo_price = spot * 0.65
    hi_price = spot * 1.35
    prices   = [lo_price + (hi_price - lo_price) * i / (width-1) for i in range(width)]

    # P&L at each price
    pnls = [strangle_pnl(p, put_strike, call_strike, total_premium, qty) for p in prices]
    max_pnl  =  total_premium
    min_pnl  = min(pnls)
    pnl_range = max_pnl - min_pnl if max_pnl != min_pnl else 1

    chart_height = 12
    rows = []
    for row in range(chart_height):
        threshold = max_pnl - (row / (chart_height-1)) * pnl_range
        line = ""
        for i, (price, pnl) in enumerate(zip(prices, pnls)):
            in_profit = pnl >= 0
            is_be_lo  = abs(price - be_lo) < (hi_price-lo_price)/width*1.5
            is_be_hi  = abs(price - be_hi) < (hi_price-lo_price)/width*1.5
            is_spot   = abs(price - spot)  < (hi_price-lo_price)/width*1.5
            is_putK   = abs(price - put_strike)  < (hi_price-lo_price)/width*1.5
            is_callK  = abs(price - call_strike) < (hi_price-lo_price)/width*1.5
            at_level  = pnl >= threshold - pnl_range/(chart_height*2)

            if at_level:
                if is_be_lo or is_be_hi:
                    line += f"{YL}|{R}"
                elif is_spot:
                    line += f"{CY}◆{R}"
                elif is_putK or is_callK:
                    line += f"{GY}┊{R}"
                elif in_profit:
                    line += f"{GR}█{R}"
                else:
                    line += f"{RD}█{R}"
            else:
                line += " "
        rows.append(line)

    # Y-axis labels
    print(f"\n  {B}{WH}P&L at Expiry — Short Strangle{R}  {GY}({days}d, spot=${spot:,.0f}){R}")
    print(f"  {GY}{'─'*width}{R}")
    for row_idx, line in enumerate(rows):
        if row_idx == 0:
            label = f"{GR}+${max_pnl:>6.2f}{R}"
        elif row_idx == chart_height//2:
            label = f"  {WH}$0.00  {R}"
        elif row_idx == chart_height-1:
            label = f"{RD}${min_pnl:>7.2f}{R}"
        else:
            label = "        "
        print(f"  {label} {line}")

    print(f"  {GY}{'─'*width}{R}")

    # X-axis: price labels
    n_labels = 6
    label_row = ""
    prev_end  = 0
    for i in range(n_labels+1):
        price = lo_price + (hi_price - lo_price) * i / n_labels
        pos   = int(i * (width-1) / n_labels)
        lbl   = f"${price:,.0f}"
        pad   = max(0, pos - prev_end)
        label_row += " " * pad + lbl
        prev_end = pos + len(lbl)
    print(f"          {GY}{label_row}{R}")

    # Legend
    print(f"""
  {GR}█{R} Profit zone   {RD}█{R} Loss zone   {CY}◆{R} Current ETH   {YL}|{R} Breakeven   {GY}┊{R} Strike

  {WH}Put strike:   {YL}${put_strike:>8,.0f}{R}     {WH}Call strike:  {YL}${call_strike:>8,.0f}{R}
  {WH}Lower B/E:    {GR}${be_lo:>8,.0f}{R}     {WH}Upper B/E:    {GR}${be_hi:>8,.0f}{R}
  {WH}Total premium:{GR} ${total_premium:>7.2f}{R}     {WH}Max loss:     {RD}Unlimited{R}
  {WH}Profit range: {GR}${be_lo:,.0f} — ${be_hi:,.0f}{R}  ({(be_hi-be_lo)/spot*100:.1f}% wide)
""")

    # Probability of profit
    T   = days / 365.0
    v   = 0.80  # fallback
    p_lo = prob_otm_put(spot, put_strike, T, RISK_FREE_RATE, v)
    p_hi = prob_otm_call(spot, call_strike, T, RISK_FREE_RATE, v)
    pop  = (p_lo + p_hi - 1) * 100   # approx P(both OTM)
    pop  = max(0, pop)
    bar_width = 40
    filled = int(pop / 100 * bar_width)
    bar = f"{GR}{'█'*filled}{GY}{'░'*(bar_width-filled)}{R}"
    print(f"  {WH}Est. Prob of Profit: {bar} {GR}{pop:.0f}%{R}\n")

# ── Strangle analysis ─────────────────────────────────────────────────────────

def show_strangle_analysis(spot, iv, days):
    T   = days / 365.0
    r   = RISK_FREE_RATE
    qty = BUDGET_USD / spot   # notional ETH exposure

    hdr(f"Short Strangle Analysis  —  {days}-day expiry")
    inf("ETH Spot", f"${spot:,.2f}")
    inf("IV",       f"{iv*100:.1f}%")
    inf("Budget",   f"${BUDGET_USD:.0f}")

    sub("Strike Pairs  (Put OTM% / Call OTM%  —  symmetric)")
    print(f"\n    {'Strikes':<22}{'Put Prem':<11}{'Call Prem':<11}"
          f"{'Combined':<12}{'Yield/yr':<12}{'Lower B/E':<13}{'Upper B/E':<13}{'P(Profit)'}")
    print(f"    {'─'*100}")

    best_row = None
    for otm in OTM_LEVELS:
        Kp  = round(spot*(1-otm)/10)*10
        Kc  = round(spot*(1+otm)/10)*10
        pp  = bs_put (spot, Kp, T, r, iv) * qty
        cp  = bs_call(spot, Kc, T, r, iv) * qty
        tot = pp + cp
        yld = (tot/BUDGET_USD)*(365/days)*100
        prem_per_eth = (pp+cp)/qty
        be_lo = Kp - prem_per_eth
        be_hi = Kc + prem_per_eth
        p_lo  = prob_otm_put (spot, Kp, T, r, iv)
        p_hi  = prob_otm_call(spot, Kc, T, r, iv)
        pop   = max(0, (p_lo + p_hi - 1)*100)

        c = GR if otm == 0.15 else WH
        print(f"    {c}${Kp:>7,.0f} / ${Kc:>7,.0f}   "
              f"${pp:>7.2f}   ${cp:>7.2f}   "
              f"${tot:>8.2f}   {yld:>7.1f}%/yr   "
              f"${be_lo:>8,.0f}   ${be_hi:>8,.0f}   {pop:.0f}%{R}")
        if otm == 0.15:
            best_row = (Kp, Kc, pp, cp, tot, qty)

    print(f"\n    {GY}* Premiums estimated via Black-Scholes  |  qty ≈ ${BUDGET_USD:.0f}/spot{R}")
    warn("Short strangles have unlimited loss potential on the call side — size carefully.")

    if best_row:
        Kp, Kc, pp, cp, tot, qty2 = best_row
        print()
        resp = input(f"  {YL}Show profit zone chart for 15% OTM strangle? (y/n): {R}").strip().lower()
        if resp == "y":
            draw_profit_zone(spot, Kp, Kc, tot, qty2, days, iv)

# ── Strangle paper trading ────────────────────────────────────────────────────

STRANGLE_STATE_FILE = "strangle_state.json"

def load_strangle():
    if os.path.exists(STRANGLE_STATE_FILE):
        with open(STRANGLE_STATE_FILE) as f:
            return json.load(f)
    return {"open":None,"total_premium":0.0,"wins":0,"losses":0,"trades":0}

def save_strangle(s):
    with open(STRANGLE_STATE_FILE,"w") as f:
        json.dump(s, f, indent=2)

def strangle_paper_menu(spot, iv, wb, days):
    global STOP_LOSS_MULTIPLIER, STOP_WARN_MULTIPLIER
    T   = days / 365.0
    s   = load_strangle()
    qty = BUDGET_USD / spot

    hdr("Short Strangle — Paper Trading")

    total = s["wins"] + s["losses"]
    inf("Trades completed",   str(total))
    inf("Wins / Losses",      f"{s['wins']} / {s['losses']}")
    inf("Win Rate",           f"{s['wins']/total*100:.1f}%" if total else "N/A")
    inf("Total Premium",      f"${s['total_premium']:.2f}")

    if s["open"]:
        op     = s["open"]
        Kp     = op["put_strike"]
        Kc     = op["call_strike"]
        p0     = op["total_premium"]
        q      = op["qty"]
        cur_pp = bs_put (spot, Kp, T, RISK_FREE_RATE, iv) * q
        cur_cp = bs_call(spot, Kc, T, RISK_FREE_RATE, iv) * q
        cur_val= cur_pp + cur_cp
        unreal = p0 - cur_val
        colour = GR if unreal >= 0 else RD

        sub("Open Strangle Position")
        inf("  Put Strike",       f"${Kp:,.0f}")
        inf("  Call Strike",      f"${Kc:,.0f}")
        inf("  Expiry",           op.get("expiry",""))
        inf("  Premium Received", f"${p0:.2f}")
        inf("  Current Value",    f"${cur_val:.2f}")
        inf("  Unrealised P&L",   f"{colour}${unreal:.2f}{R}")

        # Breakevens
        prem_per_eth = p0 / q
        be_lo = Kp - prem_per_eth
        be_hi = Kc + prem_per_eth
        inf("  Lower breakeven",  f"${be_lo:,.0f}")
        inf("  Upper breakeven",  f"${be_hi:,.0f}")
        inf("  ETH needs to stay",f"${be_lo:,.0f} — ${be_hi:,.0f}")

        # Stop-loss status
        sl_status, sl_val, sl_mult, sl_msg = check_stop_loss(spot, iv, days, op)
        print_stop_loss_status(sl_status, sl_val, sl_mult, sl_msg, p0)

    print(f"""
  {CY}[1]{R} Open new strangle
  {CY}[2]{R} Show profit zone chart (open position)
  {CY}[3]{R} Close / expire position at expiry price
  {CY}[4]{R} Close position NOW at current price  {RD}(stop-loss / early exit){R}
  {CY}[5]{R} Adjust stop-loss multiplier  {GY}(currently {STOP_LOSS_MULTIPLIER:.1f}x){R}
  {CY}[6]{R} Back
""")
    choice = input(f"  {YL}Choice: {R}").strip()

    if choice == "1":
        if s["open"]:
            warn("Close existing position first."); return

        sub("Suggested strikes (15% OTM each side)")
        Kp_sug = round(spot*0.85/10)*10
        Kc_sug = round(spot*1.15/10)*10
        pp_sug = bs_put (spot, Kp_sug, T, RISK_FREE_RATE, iv) * qty
        cp_sug = bs_call(spot, Kc_sug, T, RISK_FREE_RATE, iv) * qty
        inf(f"  Put strike  (15% below)", f"${Kp_sug:,.0f}  → ${pp_sug:.2f} premium")
        inf(f"  Call strike (15% above)", f"${Kc_sug:,.0f}  → ${cp_sug:.2f} premium")
        inf(f"  Combined premium",        f"${pp_sug+cp_sug:.2f}")

        Kp = float(input(f"\n  Put strike  [enter for ${Kp_sug:,.0f}]: $") or Kp_sug)
        Kc = float(input(f"  Call strike [enter for ${Kc_sug:,.0f}]: $") or Kc_sug)
        pp = bs_put (spot, Kp, T, RISK_FREE_RATE, iv) * qty
        cp = bs_call(spot, Kc, T, RISK_FREE_RATE, iv) * qty
        tot = pp + cp
        expiry = (date.today() + timedelta(days=days)).strftime("%d-%b-%Y")

        s["open"] = {"put_strike":Kp,"call_strike":Kc,"total_premium":round(tot,4),
                     "qty":qty,"expiry":expiry,"spot_open":spot,"days":days}
        s["total_premium"] += tot
        s["trades"] += 1
        save_strangle(s)
        ok(f"Strangle opened: Put ${Kp:,.0f} / Call ${Kc:,.0f}  |  Premium: ${tot:.2f}")

        draw_profit_zone(spot, Kp, Kc, tot, qty, days, iv)

        trade = {"date":str(date.today()),"type":"Short Strangle — Open",
                 "put_strike":Kp,"call_strike":Kc,"spot_open":spot,
                 "premium":round(tot,4),"days":days,"result":"Open",
                 "notes":f"Paper strangle, {days}d expiry"}
        append_strangle_row(wb, trade)

    elif choice == "2":
        if not s["open"]:
            warn("No open position."); return
        op  = s["open"]
        tot = op["total_premium"]
        q   = op["qty"]
        draw_profit_zone(spot, op["put_strike"], op["call_strike"], tot, q, days, iv)

    elif choice == "3":
        # Close at expiry price
        if not s["open"]:
            warn("No open position."); return
        op = s["open"]
        Kp = op["put_strike"]; Kc = op["call_strike"]
        tot= op["total_premium"]; q = op["qty"]
        spot_close = float(input(f"  ETH price at expiry [~${spot:,.0f}]: $") or spot)
        pnl = strangle_pnl(spot_close, Kp, Kc, tot, q)
        if pnl >= 0:
            result = "Win"; s["wins"] += 1
            ok(f"Expired in profit!  P&L: +${pnl:.2f}")
        else:
            result = "Loss"; s["losses"] += 1
            side = "PUT side (ETH dropped)" if spot_close < Kp else "CALL side (ETH rallied)"
            warn(f"Expired in loss.  P&L: ${pnl:.2f}  ({side})")
        trade = {"date":str(date.today()),"type":"Short Strangle — Expired",
                 "put_strike":Kp,"call_strike":Kc,
                 "spot_open":op["spot_open"],"spot_close":spot_close,
                 "premium":round(tot,4),"pnl":round(pnl,4),
                 "days":op["days"],"result":result,"notes":"Held to expiry"}
        append_strangle_row(wb, trade)
        s["open"] = None; save_strangle(s)

    elif choice == "4":
        # Close NOW at current market price (stop-loss / early exit)
        if not s["open"]:
            warn("No open position."); return
        op  = s["open"]
        Kp  = op["put_strike"]; Kc = op["call_strike"]
        p0  = op["total_premium"]; q = op["qty"]
        sl_status, cur_val, mult, sl_msg = check_stop_loss(spot, iv, days, op)
        pnl = p0 - cur_val
        sub("Early Close — Mark-to-Market")
        inf("  Cost to buy back now",         f"${cur_val:.2f}  ({mult:.2f}x premium)")
        inf("  Premium originally collected", f"${p0:.2f}")
        colour2 = GR if pnl >= 0 else RD
        inf("  Net P&L if closed now",         f"{colour2}${pnl:.2f}{R}")
        if sl_status == "stop":
            print(f"\n  {RD}⛔  Stop-loss level reached — strongly recommended to close.{R}")
        elif sl_status == "warn":
            print(f"\n  {YL}⚠  Warning level — consider closing before it gets worse.{R}")
        else:
            print(f"\n  {GY}Position within normal range — early close is optional.{R}")
        confirm = input(f"\n  {YL}Confirm close at current price? (y/n): {R}").strip().lower()
        if confirm != "y":
            print("  Cancelled."); return
        result = "Win" if pnl >= 0 else ("Loss (Stop)" if sl_status=="stop" else "Loss (Early)")
        if pnl >= 0: s["wins"] += 1
        else:        s["losses"] += 1
        note = (f"Stop-loss at {mult:.1f}x — loss ${abs(pnl):.2f}" if sl_status=="stop"
                else f"Early close at {mult:.1f}x — P&L ${pnl:.2f}")
        if pnl >= 0: ok(f"Closed.  P&L: +${pnl:.2f}")
        else:        warn(f"Closed.  P&L: ${pnl:.2f}")
        trade = {"date":str(date.today()),"type":"Short Strangle — Early Close",
                 "put_strike":Kp,"call_strike":Kc,
                 "spot_open":op["spot_open"],"spot_close":spot,
                 "premium":round(p0,4),"pnl":round(pnl,4),
                 "days":op["days"],"result":result,"notes":note}
        append_strangle_row(wb, trade)
        s["open"] = None; save_strangle(s)

    elif choice == "5":
        # Adjust stop-loss multipliers
        sub("Adjust Stop-Loss Settings")
        inf("Current stop multiplier", f"{STOP_LOSS_MULTIPLIER:.1f}x")
        inf("Current warn multiplier", f"{STOP_WARN_MULTIPLIER:.1f}x")
        print(f"""
  {GY}Common rules used by options sellers:{R}
  {CY}2x{R} stop = lose at most 1x your premium  {GY}(conservative — recommended for beginners){R}
  {CY}3x{R} stop = lose at most 2x your premium  {GY}(moderate){R}
  {CY}5x{R} stop = lose at most 4x your premium  {GY}(aggressive — not recommended){R}
""")
        new_stop = input(f"  New stop multiplier [{STOP_LOSS_MULTIPLIER:.1f}]: ").strip()
        new_warn = input(f"  New warn multiplier [{STOP_WARN_MULTIPLIER:.1f}]: ").strip()
        if new_stop: STOP_LOSS_MULTIPLIER = float(new_stop)
        if new_warn: STOP_WARN_MULTIPLIER = float(new_warn)
        ok(f"Updated: warn at {STOP_WARN_MULTIPLIER:.1f}x, stop at {STOP_LOSS_MULTIPLIER:.1f}x")
        if STOP_LOSS_MULTIPLIER > 3.0:
            warn("Stop > 3x is aggressive. One bad crypto move can wipe weeks of gains.")

    else:
        return

# ── Excel helpers ─────────────────────────────────────────────────────────────

DARK   = "1A1A2E"; ACCENT = "E94560"; GOLD = "F5A623"
MID    = "16213E"; LIGHT  = "E8E8E8"; GREEN_C = "27AE60"; RED_C = "E74C3C"
WHITE_C= "FFFFFF"

def _thin_border():
    t = Side(style="thin", color="444444")
    return Border(left=t, right=t, top=t, bottom=t)

def _styled(ws, ref, val, bold=False, bg=None, fg="000000",
            align="center", num_fmt=None, italic=False, size=10):
    c = ws[ref] if isinstance(ref, str) else ref
    c.value = val
    c.font  = Font(name="Arial", bold=bold, italic=italic, color=fg, size=size)
    if bg: c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin_border()
    if num_fmt: c.number_format = num_fmt
    return c


def setup_excel():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        # Add strangle sheet if missing
        if "🔀 Strangles" not in wb.sheetnames:
            _create_strangle_sheet(wb)
            wb.save(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _create_dashboard(wb)
        _create_trade_sheet(wb, "📋 Live Trades")
        _create_trade_sheet(wb, "📝 Paper Trades")
        _create_strangle_sheet(wb)
        _create_summary_sheet(wb)
        wb.save(EXCEL_FILE)
    return wb


def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _create_dashboard(wb):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3,32,22,22,22])

    ws.row_dimensions[2].height = 45
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "ETH OPTIONS STRATEGY TRACKER  v3"
    c.font  = Font(name="Arial", bold=True, color=ACCENT, size=20)
    c.fill  = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        (4,  "Budget (USD)",                    "250",                              "$#,##0.00", GOLD),
        (5,  "Wheel — Total Paper Premium",     "='📝 Paper Trades'!K2",           "$#,##0.00", GREEN_C),
        (6,  "Wheel — Paper Win Rate",          "='📝 Paper Trades'!L2",           "0.0%",      GOLD),
        (7,  "Wheel — Paper Cycles",            "='📝 Paper Trades'!M2",           "0",         LIGHT),
        (8,  "Strangle — Total Paper Premium",  "='🔀 Strangles'!H2",              "$#,##0.00", GREEN_C),
        (9,  "Strangle — Paper Win Rate",       "='🔀 Strangles'!I2",              "0.0%",      GOLD),
        (10, "Strangle — Trades Completed",     "='🔀 Strangles'!J2",              "0",         LIGHT),
        (11, "Live — Total Premium",            "='📋 Live Trades'!K2",            "$#,##0.00", GREEN_C),
        (12, "Live — Win Rate",                 "='📋 Live Trades'!L2",            "0.0%",      GOLD),
    ]
    for row, label, formula, fmt, color in kpis:
        ws.row_dimensions[row].height = 24
        _styled(ws, f"B{row}", label, bold=True, bg=MID, fg=LIGHT, align="left")
        _styled(ws, f"C{row}", formula, bg=DARK, fg=color, num_fmt=fmt)

    ws.row_dimensions[14].height = 18
    ws.merge_cells("B14:E14")
    c = ws["B14"]
    c.value = "ℹ  Premiums are Black-Scholes estimates. Paper trading only — not financial advice."
    c.font  = Font(name="Arial", italic=True, color="888888", size=9)
    c.alignment = Alignment(horizontal="center")

    # Strategy comparison
    ws.row_dimensions[16].height = 22
    ws.merge_cells("B16:E16")
    _styled(ws, "B16", "STRATEGY COMPARISON", bold=True, bg=MID, fg=GOLD, size=11)
    hdrs = ["Strategy","Max Income","Risk","Best Market"]
    for col, h in enumerate(hdrs, 2):
        _styled(ws, f"{get_column_letter(col)}17", h, bold=True, bg=DARK, fg=ACCENT)
    rows_data = [
        ("Wheel (CSP→CC)", "Medium", "Capped (put side)", "Mild bullish / sideways"),
        ("Short Strangle",  "High",  "Unlimited",         "Sideways / low vol"),
        ("Short Straddle",  "Highest","Unlimited",        "Very sideways"),
        ("Iron Condor",     "Low",   "Capped ✓",          "Range-bound"),
    ]
    for i, row_data in enumerate(rows_data, 18):
        ws.row_dimensions[i].height = 20
        for col, val in enumerate(row_data, 2):
            _styled(ws, f"{get_column_letter(col)}{i}", val,
                    bg=MID if i%2==0 else DARK, fg=LIGHT, align="left")


def _create_trade_sheet(wb, name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3,13,22,18,10,13,14,14,13,13,10,18,12,12,12,28])

    ws.row_dimensions[2].height = 1
    ws["K2"] = "=IF(COUNTA(I4:I5000)=0,0,SUM(I4:I5000))"
    ws["L2"] = '=IF(COUNTA(J4:J5000)=0,0,COUNTIF(J4:J5000,"Win")/COUNTA(J4:J5000))'
    ws["M2"] = '=COUNTIF(C4:C5000,"Call Expired Worthless")'
    ws["K2"].number_format = "$#,##0.00"
    ws["L2"].number_format = "0.0%"

    ws.row_dimensions[3].height = 34
    headers = ["Date","Type","Stage","Days","Strike ($)","Spot Open ($)","Spot Close ($)",
               "Premium ($)","P&L ($)","Result","Cum. Premium ($)","Win Rate","Cycles","","Notes"]
    for col, h in enumerate(headers, 2):
        ref = f"{get_column_letter(col)}3"
        ws[ref] = h
        ws[ref].font = Font(name="Arial", bold=True, color=ACCENT, size=10)
        ws[ref].fill = PatternFill("solid", start_color=DARK)
        ws[ref].border = _thin_border()
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 18
    c = ws["B4"]
    c.value = "← Trades will appear here as you use the tool"
    c.font  = Font(name="Arial", italic=True, color="666666", size=9)
    c.fill  = PatternFill("solid", start_color=MID)
    ws.merge_cells("B4:P4")
    c.alignment = Alignment(horizontal="center", vertical="center")


def _create_strangle_sheet(wb):
    ws = wb.create_sheet("🔀 Strangles")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3,13,14,14,14,14,14,14,12,12,16,16,28])

    ws.row_dimensions[2].height = 1
    ws["H2"] = '=IF(COUNTA(H4:H5000)=0,0,SUM(H4:H5000))'
    ws["I2"] = '=IF(COUNTA(K4:K5000)=0,0,COUNTIF(K4:K5000,"Win")/COUNTA(K4:K5000))'
    ws["J2"] = '=COUNTA(A4:A5000)'
    ws["H2"].number_format = "$#,##0.00"
    ws["I2"].number_format = "0.0%"

    ws.row_dimensions[3].height = 34
    headers = ["Date","Type","Put Strike","Call Strike","Spot Open","Spot Close",
               "Days","Premium ($)","P&L ($)","Lower B/E","Upper B/E","Result","Notes"]
    for col, h in enumerate(headers, 2):
        ref = f"{get_column_letter(col)}3"
        ws[ref] = h
        ws[ref].font = Font(name="Arial", bold=True, color=ACCENT, size=10)
        ws[ref].fill = PatternFill("solid", start_color=DARK)
        ws[ref].border = _thin_border()
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 18
    c = ws["B4"]
    c.value = "← Strangle trades will appear here as you paper trade"
    c.font  = Font(name="Arial", italic=True, color="666666", size=9)
    c.fill  = PatternFill("solid", start_color=MID)
    ws.merge_cells("B4:N4")
    c.alignment = Alignment(horizontal="center", vertical="center")


def _create_summary_sheet(wb):
    ws = wb.create_sheet("📈 Summary")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [3,18,16,16,16,16,30])
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "CYCLE PERFORMANCE SUMMARY"
    c.font  = Font(name="Arial", bold=True, color=ACCENT, size=16)
    c.fill  = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

def _unmerge_row(ws, row, max_col=17):
    """Unmerge any merged cells in this row so we can write to them freely."""
    merges_to_remove = [
        rng for rng in ws.merged_cells.ranges
        if rng.min_row <= row <= rng.max_row
    ]
    for rng in merges_to_remove:
        ws.unmerge_cells(str(rng))
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).value = None

def append_trade_row(wb, sheet_name, trade):
    ws = wb[sheet_name]
    row = 4
    while True:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            break
        if isinstance(cell_val, str) and "←" in cell_val:
            _unmerge_row(ws, row)   # unmerge & clear the placeholder row
            break
        row += 1

    result  = trade.get("result","")
    bg      = "1a3a1a" if result=="Win" else "3a1a1a" if result=="Loss" else MID
    fg      = GREEN_C  if result=="Win" else RED_C    if result=="Loss" else LIGHT

    values = [trade.get("date",""), trade.get("type",""), trade.get("stage",""),
              trade.get("days",7), trade.get("strike",0), trade.get("spot_open",0),
              trade.get("spot_close",""), trade.get("premium",0),
              trade.get("pnl",""), result,
              f"=SUM($I$4:I{row})", "", "", "", trade.get("notes","")]

    fmts = {5:"$#,##0",6:"$#,##0.00",7:"$#,##0.00",8:"$#,##0.00",9:"$#,##0.00",11:"$#,##0.00"}
    for col_idx, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font   = Font(name="Arial", size=9, color=fg)
        c.fill   = PatternFill("solid", start_color=bg)
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in fmts: c.number_format = fmts[col_idx]
    ws.row_dimensions[row].height = 18
    wb.save(EXCEL_FILE)
    ok(f"Trade saved → {sheet_name} row {row}")


def append_strangle_row(wb, trade):
    ws = wb["🔀 Strangles"]
    row = 4
    while True:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            break
        if isinstance(cell_val, str) and "←" in cell_val:
            _unmerge_row(ws, row, max_col=14)
            break
        row += 1

    result = trade.get("result","")
    bg     = "1a3a1a" if result=="Win" else "3a1a1a" if result=="Loss" else MID
    fg     = GREEN_C  if result=="Win" else RED_C    if result=="Loss" else LIGHT

    Kp   = trade.get("put_strike",0)
    Kc   = trade.get("call_strike",0)
    prem = trade.get("premium",0)
    qty  = BUDGET_USD / trade.get("spot_open",1) if trade.get("spot_open") else 0
    prem_per_eth = prem / qty if qty else 0
    be_lo = Kp - prem_per_eth
    be_hi = Kc + prem_per_eth

    values = [trade.get("date",""), trade.get("type",""),
              Kp, Kc, trade.get("spot_open",""), trade.get("spot_close",""),
              trade.get("days",7), prem,
              trade.get("pnl",""), be_lo, be_hi, result, trade.get("notes","")]

    fmts = {4:"$#,##0",5:"$#,##0",6:"$#,##0.00",7:"$#,##0.00",
            8:"$#,##0.00",9:"$#,##0.00",10:"$#,##0",11:"$#,##0"}
    for col_idx, val in enumerate(values, 2):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font   = Font(name="Arial", size=9, color=fg)
        c.fill   = PatternFill("solid", start_color=bg)
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in fmts: c.number_format = fmts[col_idx]
    ws.row_dimensions[row].height = 18
    wb.save(EXCEL_FILE)
    ok(f"Strangle trade saved → 🔀 Strangles row {row}")

# ── Wheel paper trading ───────────────────────────────────────────────────────

PAPER_STATE_FILE = "paper_state.json"

def load_paper():
    if os.path.exists(PAPER_STATE_FILE):
        with open(PAPER_STATE_FILE) as f:
            return json.load(f)
    return {"stage":"no_position","open":None,"eth_held":0.0,
            "cost_basis":0.0,"total_premium":0.0,
            "wins":0,"losses":0,"cycles":0}

def save_paper(s):
    with open(PAPER_STATE_FILE,"w") as f:
        json.dump(s, f, indent=2)

def wheel_paper_menu(spot, iv, wb, days):
    T = days / 365.0
    s = load_paper()

    hdr("Wheel Strategy — Paper Trading")
    stage_labels = {
        "no_position": "No Position — ready to sell a Put",
        "short_put":   "Short Put open — waiting for expiry",
        "holding_eth": "Holding ETH — ready to sell a Call",
        "short_call":  "Short Call open — waiting for expiry",
    }
    inf("Current Stage",    stage_labels.get(s["stage"],"?"))
    inf("Total Premium",    f"${s['total_premium']:.2f}")
    total = s["wins"]+s["losses"]
    inf("Wins / Losses",    f"{s['wins']} / {s['losses']}")
    inf("Win Rate",         f"{s['wins']/total*100:.1f}%" if total else "N/A")
    inf("Cycles Completed", str(s["cycles"]))

    if s["open"]:
        op = s["open"]; K = op["strike"]; p0 = op["premium"]
        cur = (bs_put(spot,K,T,RISK_FREE_RATE,iv) if op["type"]=="Put"
               else bs_call(spot,K,T,RISK_FREE_RATE,iv)) * (BUDGET_USD/K)
        unreal = p0 - cur
        sub("Open Position")
        inf("  Type",             op["type"])
        inf("  Strike",           f"${K:,.0f}")
        inf("  Expiry",           op.get("expiry",""))
        inf("  Premium Received", f"${p0:.2f}")
        colour = GR if unreal >= 0 else RD
        inf("  Unrealised P&L",   f"{colour}${unreal:.2f}{R}")

    print(f"""
  {CY}[1]{R} Sell Put (open position)
  {CY}[2]{R} Expire position (enter ETH price at expiry)
  {CY}[3]{R} Assign put (take ETH)
  {CY}[4]{R} Sell Covered Call
  {CY}[5]{R} Back
""")
    choice = input(f"  {YL}Choice: {R}").strip()

    if choice == "1":
        if s["stage"] != "no_position": warn("Close existing position first."); return
        K_sug = round(spot*0.85/10)*10
        p_sug = bs_put(spot, K_sug, T, RISK_FREE_RATE, iv) * (BUDGET_USD/K_sug)
        inf(f"Suggested (15% OTM)", f"${K_sug:,.0f}  → ${p_sug:.2f} premium")
        K       = float(input(f"  Strike [enter for ${K_sug:,.0f}]: $") or K_sug)
        qty     = BUDGET_USD / K
        premium = bs_put(spot, K, T, RISK_FREE_RATE, iv) * qty
        expiry  = (date.today() + timedelta(days=days)).strftime("%d-%b-%Y")
        s["stage"]="short_put"
        s["open"]={"type":"Put","strike":K,"expiry":expiry,
                   "premium":round(premium,4),"spot_open":spot,"qty":qty,"days":days}
        s["total_premium"]+=premium
        save_paper(s)
        ok(f"Sell Put @ ${K:,.0f}  |  Premium: ${premium:.2f}")
        append_trade_row(wb,"📝 Paper Trades",
            {"date":str(date.today()),"type":"Sell Cash-Secured Put","stage":"Short Put",
             "days":days,"strike":K,"spot_open":spot,"premium":round(premium,4),
             "result":"Open","notes":f"Paper, {days}d"})

    elif choice == "2":
        if not s["open"]: warn("No open position."); return
        op=s["open"]; K=op["strike"]; p0=op["premium"]
        spot_close=float(input(f"  ETH at expiry [~${spot:,.0f}]: $") or spot)
        expired = spot_close>K if op["type"]=="Put" else spot_close<K
        if expired:
            pnl=p0; result="Win"; s["wins"]+=1
            ok(f"Expired worthless ✓  P&L: +${pnl:.2f}")
            s["stage"]="holding_eth" if op["type"]=="Call" else "no_position"
            if op["type"]=="Call": s["cycles"]+=1
        else:
            intrinsic=abs(spot_close-K)*op.get("qty",BUDGET_USD/K)
            pnl=p0-intrinsic; result="Loss"; s["losses"]+=1
            warn(f"ITM loss  P&L: ${pnl:.2f}"); s["stage"]="no_position"
        append_trade_row(wb,"📝 Paper Trades",
            {**op,"date":str(date.today()),
             "type":f"{'Put' if op['type']=='Put' else 'Call'}→Expired",
             "stage":"Closed","spot_close":spot_close,
             "pnl":round(pnl,4),"result":result,"notes":f"Paper expired {days}d"})
        s["open"]=None; save_paper(s)

    elif choice == "3":
        if s["stage"]!="short_put": warn("Not in short put stage."); return
        op=s["open"]
        s["eth_held"]=op.get("qty",BUDGET_USD/op["strike"])
        s["cost_basis"]=op["strike"]; s["stage"]="holding_eth"; s["open"]=None
        save_paper(s)
        ok(f"Assigned! {s['eth_held']:.4f} ETH @ ${op['strike']:,.0f}")

    elif choice == "4":
        if s["stage"]!="holding_eth": warn("Hold ETH first."); return
        K_sug=round(spot*1.15/10)*10
        p_sug=bs_call(spot,K_sug,T,RISK_FREE_RATE,iv)*s["eth_held"]
        inf(f"Suggested (15% OTM)", f"${K_sug:,.0f}  → ${p_sug:.2f} premium")
        K=float(input(f"  Strike [enter for ${K_sug:,.0f}]: $") or K_sug)
        qty=s["eth_held"]
        premium=bs_call(spot,K,T,RISK_FREE_RATE,iv)*qty
        expiry=(date.today()+timedelta(days=days)).strftime("%d-%b-%Y")
        s["stage"]="short_call"
        s["open"]={"type":"Call","strike":K,"expiry":expiry,
                   "premium":round(premium,4),"spot_open":spot,"qty":qty,"days":days}
        s["total_premium"]+=premium; save_paper(s)
        ok(f"Sell Call @ ${K:,.0f}  |  Premium: ${premium:.2f}")
        append_trade_row(wb,"📝 Paper Trades",
            {"date":str(date.today()),"type":"Sell Covered Call","stage":"Short Call",
             "days":days,"strike":K,"spot_open":spot,"premium":round(premium,4),
             "result":"Open","notes":f"Paper, {days}d"})
    else:
        return

# ── Strike analysis (wheel) ───────────────────────────────────────────────────

def show_strikes(spot, iv, days):
    T = days/365.0; r = RISK_FREE_RATE
    sub(f"Cash-Secured Put Strikes  ({days}-day)")
    print(f"\n    {'OTM%':<8}{'Strike':<12}{'Total Prem':<14}{'Yield/yr':<13}{'P(Profit)'}")
    print(f"    {'─'*55}")
    for otm in OTM_LEVELS:
        K=round(spot*(1-otm)/10)*10; qty=BUDGET_USD/K
        tot=bs_put(spot,K,T,r,iv)*qty
        yld=(tot/BUDGET_USD)*(365/days)*100
        pp=prob_otm_put(spot,K,T,r,iv)*100
        c=GR if otm==0.15 else WH
        print(f"    {c}{otm*100:.0f}%{'':<6}${K:>8,.0f}   ${tot:>7.2f}       {yld:>6.1f}%/yr   {pp:.0f}%{R}")

    sub(f"Covered Call Strikes  ({days}-day)")
    print(f"\n    {'OTM%':<8}{'Strike':<12}{'Total Prem':<14}{'Yield/yr':<13}{'P(Profit)'}")
    print(f"    {'─'*55}")
    qty=BUDGET_USD/spot
    for otm in OTM_LEVELS:
        K=round(spot*(1+otm)/10)*10
        tot=bs_call(spot,K,T,r,iv)*qty
        yld=(tot/BUDGET_USD)*(365/days)*100
        pp=prob_otm_call(spot,K,T,r,iv)*100
        c=GR if otm==0.15 else WH
        print(f"    {c}{otm*100:.0f}%{'':<6}${K:>8,.0f}   ${tot:>7.2f}       {yld:>6.1f}%/yr   {pp:.0f}%{R}")
    print(f"\n    {GY}IV: {iv*100:.0f}%  |  Budget: ${BUDGET_USD:.0f}  |  BS estimate{R}")

# ── Cycle summary ─────────────────────────────────────────────────────────────

def show_summary(wb):
    hdr("Performance Summary")
    for sheet_name in ["📝 Paper Trades","📋 Live Trades","🔀 Strangles"]:
        ws = wb[sheet_name]
        sub(sheet_name)
        rows = [r for r in ws.iter_rows(min_row=4,values_only=True)
                if r[0] and "←" not in str(r[0])]
        if not rows: inf("No trades yet",""); continue
        col_result = 9 if sheet_name != "🔀 Strangles" else 11
        col_prem   = 7 if sheet_name != "🔀 Strangles" else 7
        wins   = sum(1 for t in rows if t[col_result]=="Win")
        losses = sum(1 for t in rows if t[col_result]=="Loss")
        total  = wins+losses
        prems  = [t[col_prem] for t in rows if isinstance(t[col_prem],(int,float))]
        inf("Trades",       str(len(rows)))
        inf("Wins/Losses",  f"{wins} / {losses}")
        inf("Win Rate",     f"{wins/total*100:.1f}%" if total else "N/A")
        inf("Total Premium",f"${sum(prems):.2f}" if prems else "$0")
        inf("Avg Premium",  f"${sum(prems)/len(prems):.2f}" if prems else "N/A")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    hdr("ETH Options Strategy Tool  v3.0  🔄")
    print(f"  {GY}Budget: ${BUDGET_USD:.0f}  |  Strategies: Wheel + Short Strangle{R}")

    sub("Fetching live market data...")
    spot = get_eth_price()
    if spot: ok(f"ETH/USD: ${spot:,.2f}")
    else:    spot = float(input(f"  {YL}Enter ETH price: ${R}") or 1600)

    days = 7
    iv   = get_deribit_iv(spot, days)
    if iv: ok(f"IV (ATM {days}d Deribit): {iv*100:.1f}%")
    else:  iv=0.80; inf("Using default IV",f"{iv*100:.0f}%")

    wb = setup_excel()
    ok(f"Excel tracker: {EXCEL_FILE}")

    while True:
        hdr("Main Menu")
        inf("ETH Spot", f"${spot:,.2f}")
        inf("IV",       f"{iv*100:.1f}%")
        inf("Expiry",   f"{days}-day")

        # Passive stop-loss alert if a strangle is open
        ss = load_strangle()
        if ss.get("open"):
            sl_status, sl_val, sl_mult, sl_msg = check_stop_loss(spot, iv, days, ss["open"])
            if sl_status == "stop":
                print(f"\n  {RD}{'━'*50}{R}")
                print(f"  {RD}  ⛔  STRANGLE STOP-LOSS TRIGGERED  ({sl_mult:.1f}x){R}")
                print(f"  {RD}  Go to [4] Strangle paper trading → [4] Close NOW{R}")
                print(f"  {RD}{'━'*50}{R}")
            elif sl_status == "warn":
                print(f"\n  {YL}  ⚠  Strangle warning: {sl_mult:.1f}x premium — approaching stop ({STOP_LOSS_MULTIPLIER:.1f}x){R}")

        print(f"""
  ── Wheel Strategy ──────────────────────────
  {CY}[1]{R} Wheel strike & premium analysis
  {CY}[2]{R} Wheel paper trading simulator

  ── Short Strangle ───────────────────────────
  {CY}[3]{R} Strangle analysis + profit zone chart
  {CY}[4]{R} Strangle paper trading simulator

  ── General ──────────────────────────────────
  {CY}[5]{R} Record live trade (wheel)
  {CY}[6]{R} Performance summary & stats
  {CY}[7]{R} Switch expiry  (currently {YL}{days}{R}-day)
  {CY}[8]{R} Refresh market data
  {CY}[9]{R} Exit
""")
        choice = input(f"  {YL}Choice: {R}").strip()

        if   choice=="1": show_strikes(spot, iv, days)
        elif choice=="2": wheel_paper_menu(spot, iv, wb, days)
        elif choice=="3": show_strangle_analysis(spot, iv, days)
        elif choice=="4": strangle_paper_menu(spot, iv, wb, days)
        elif choice=="5":
            # quick live wheel trade entry
            hdr("Record Live Wheel Trade")
            type_map={"1":("Sell CSP","Short Put","Open"),"2":("Put Expired","Closed","Win"),
                      "3":("Put Assigned","Holding ETH",""),"4":("Sell Call","Short Call","Open"),
                      "5":("Call Expired","Closed","Win"),"6":("Call Assigned","Closed","Win")}
            for k,(t,_,_) in type_map.items():
                print(f"  {CY}[{k}]{R} {t}")
            c2=input(f"\n  {YL}Choice: {R}").strip()
            if c2 in type_map:
                t,stage,result=type_map[c2]
                K=float(input("  Strike $") or 0)
                sp=float(input("  Spot at open $") or spot)
                pr=float(input("  Premium received $") or 0)
                no=input("  Notes: ")
                append_trade_row(wb,"📋 Live Trades",
                    {"date":str(date.today()),"type":t,"stage":stage,
                     "days":days,"strike":K,"spot_open":sp,"premium":pr,
                     "result":result,"notes":no})
        elif choice=="6": show_summary(wb)
        elif choice=="7":
            print(f"  {CY}[1]{R} Daily  {CY}[2]{R} Weekly")
            ec=input(f"  {YL}Choice: {R}").strip()
            days=1 if ec=="1" else 7
            iv=get_deribit_iv(spot,days) or iv
            ok(f"Switched to {days}-day  |  IV: {iv*100:.1f}%")
            if days==1: warn("Daily expiry — paper trading practice only.")
        elif choice=="8":
            ns=get_eth_price()
            if ns: spot=ns; ok(f"ETH: ${spot:,.2f}")
            ni=get_deribit_iv(spot,days)
            if ni: iv=ni; ok(f"IV: {iv*100:.1f}%")
            # Re-check stop loss after price refresh
            ss2 = load_strangle()
            if ss2.get("open"):
                sl2, sv2, sm2, smsg2 = check_stop_loss(spot, iv, days, ss2["open"])
                print_stop_loss_status(sl2, sv2, sm2, smsg2, ss2["open"]["total_premium"])
        elif choice=="9":
            print(f"\n  {GR}Good luck! Manage your risk. 👋{R}\n"); break
        else:
            err("Invalid choice.")

if __name__ == "__main__":
    main()
